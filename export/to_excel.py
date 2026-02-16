"""Strategic market intelligence report for UET Construction.

This is not a data dump — it's a narrative-driven intelligence brief
designed for a client meeting with UET's leadership.
"""

from __future__ import annotations

from decimal import Decimal
from typing import Any

import asyncpg
import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analysis.market_intel import (
    get_big_tashkent_tenders,
    get_market_summary,
    get_market_summary_12m,
    get_monthly_trend,
    get_peer_comparison,
    get_peer_rating_comparison,
    get_rating_distribution,
    get_regional_distribution,
    get_tashkent_competitors,
    get_tashkent_customers,
    get_top_companies_overall,
    get_top_customers,
    get_top50_benchmark,
    get_uet_competitiveness_gaps,
    get_uet_profile,
    get_uet_rating_breakdown,
    get_uet_rating_detail,
    get_uet_rating_percentile,
)

# ── Colour palette ──────────────────────────────────────────

# Primary palette
NAVY = "1F4E79"
WHITE = "FFFFFF"
LIGHT_BLUE = "D6E4F0"
ALT_ROW = "F2F7FB"

# Semantic colours
GREEN_BG = "E2EFDA"     # strengths
GREEN_FONT = "375623"
RED_BG = "FCE4EC"        # gaps / weaknesses
RED_FONT = "9C2731"
YELLOW_BG = "FFF8E1"     # opportunities
YELLOW_FONT = "7F6003"
UET_HIGHLIGHT = "BDD7EE"  # UET row highlight

# Rating scale
RATING_COLORS: dict[str, str] = {
    "AAA": "1B5E20", "AA": "2E7D32", "A": "388E3C",
    "BBB": "558B2F", "BB": "689F38", "B": "7CB342",
    "CCC": "F9A825", "CC": "F57F17", "C": "EF6C00",
    "DDD": "D84315", "DD": "C62828", "D": "B71C1C",
}

# ── Styles ──────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color=NAVY)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
SMALL_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
CURRENCY_FORMAT = '#,##0'
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")

# Callout box styles
GREEN_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
RED_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
YELLOW_FILL = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
UET_ROW_FILL = PatternFill(start_color=UET_HIGHLIGHT, end_color=UET_HIGHLIGHT, fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")


def _fmt_uzs(value: Any) -> str:
    """Format number as UZS with billions/trillions label."""
    if value is None:
        return "—"
    try:
        v = float(value)
    except (ValueError, TypeError):
        return str(value)
    if abs(v) >= 1_000_000_000_000:
        return f"{v / 1_000_000_000_000:,.2f} трлн UZS"
    if abs(v) >= 1_000_000_000:
        return f"{v / 1_000_000_000:,.1f} млрд UZS"
    if abs(v) >= 1_000_000:
        return f"{v / 1_000_000:,.0f} млн UZS"
    return f"{v:,.0f} UZS"


class ExcelReportGenerator:
    """Generates a story-driven strategic intelligence report."""

    def __init__(self, pool: asyncpg.Pool) -> None:
        self.pool = pool

    async def generate_full_report(
        self,
        output_path: str,
        uet_stir: str,
        compare_stirs: list[str] | None = None,
    ) -> str:
        """Generate the complete 5-sheet intelligence report."""
        wb = Workbook()

        logger.info("Sheet 1/5: Обзор рынка")
        await self._sheet_market_overview(wb)

        logger.info("Sheet 2/5: Топ-15 компаний")
        await self._sheet_top_companies(wb, uet_stir)

        logger.info("Sheet 3/5: Профиль UET")
        await self._sheet_uet_profile(wb, uet_stir)

        logger.info("Sheet 4/5: Конкуренты")
        await self._sheet_competitors(wb, uet_stir, compare_stirs)

        logger.info("Sheet 5/5: Рекомендации")
        await self._sheet_recommendations(wb, uet_stir)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        logger.info("Report saved to {}", output_path)
        return output_path

    # ═══════════════════════════════════════════════════════════
    #  SHEET 1: ОБЗОР РЫНКА (Market Overview)
    # ═══════════════════════════════════════════════════════════

    async def _sheet_market_overview(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Обзор рынка")

        mkt_all = await get_market_summary(self.pool)
        mkt_12m = await get_market_summary_12m(self.pool)
        monthly = await get_monthly_trend(self.pool)
        regional = await get_regional_distribution(self.pool)
        top_cust = await get_top_customers(self.pool, limit=10)
        rating_dist = await get_rating_distribution(self.pool)

        # ── Title ──
        ws.merge_cells("A1:H1")
        ws["A1"] = "РЫНОК СТРОИТЕЛЬНЫХ ТЕНДЕРОВ УЗБЕКИСТАНА"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

        ws.merge_cells("A2:H2")
        ws["A2"] = "Аналитический обзор на основе данных etender.uzex.uz и reyting.mc.uz"
        ws["A2"].font = SMALL_FONT

        # ── Key metrics callout boxes (row 4-5) ──
        metrics = [
            ("A", "Объём рынка (всего)", _fmt_uzs(mkt_all.get("total_volume", 0))),
            ("C", "Объём за 12 мес.", _fmt_uzs(mkt_12m.get("total_volume", 0))),
            ("E", "Тендеров (12 мес.)", f"{mkt_12m.get('total_tenders', 0):,}"),
            ("G", "Средняя скидка", f"{mkt_12m.get('avg_discount', 0)}%"),
        ]
        for col_letter, label, value in metrics:
            col_idx = ord(col_letter) - ord("A") + 1
            # Merge 2 columns for each box
            ws.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 1)
            ws.merge_cells(start_row=5, start_column=col_idx, end_row=5, end_column=col_idx + 1)
            cell_label = ws.cell(row=4, column=col_idx, value=label)
            cell_label.font = Font(name="Calibri", size=9, color="666666")
            cell_label.fill = LIGHT_BLUE_FILL
            cell_label.alignment = Alignment(horizontal="center")
            ws.cell(row=4, column=col_idx + 1).fill = LIGHT_BLUE_FILL
            cell_value = ws.cell(row=5, column=col_idx, value=value)
            cell_value.font = Font(name="Calibri", size=13, bold=True, color=NAVY)
            cell_value.fill = LIGHT_BLUE_FILL
            cell_value.alignment = Alignment(horizontal="center")
            ws.cell(row=5, column=col_idx + 1).fill = LIGHT_BLUE_FILL

        # Second row of metrics
        metrics2 = [
            ("A", "Активных компаний", f"{mkt_12m.get('unique_winners', 0):,}"),
            ("C", "Ср. размер контракта", _fmt_uzs(mkt_12m.get("avg_deal_size", 0))),
            ("E", "Ср. участников", f"{mkt_12m.get('avg_participants', 0)}"),
            ("G", "Компаний с рейтингом", f"{len(rating_dist):,} категорий"),
        ]
        for col_letter, label, value in metrics2:
            col_idx = ord(col_letter) - ord("A") + 1
            ws.merge_cells(start_row=6, start_column=col_idx, end_row=6, end_column=col_idx + 1)
            ws.merge_cells(start_row=7, start_column=col_idx, end_row=7, end_column=col_idx + 1)
            c1 = ws.cell(row=6, column=col_idx, value=label)
            c1.font = Font(name="Calibri", size=9, color="666666")
            c1.fill = LIGHT_BLUE_FILL
            c1.alignment = Alignment(horizontal="center")
            ws.cell(row=6, column=col_idx + 1).fill = LIGHT_BLUE_FILL
            c2 = ws.cell(row=7, column=col_idx, value=value)
            c2.font = Font(name="Calibri", size=12, bold=True, color=NAVY)
            c2.fill = LIGHT_BLUE_FILL
            c2.alignment = Alignment(horizontal="center")
            ws.cell(row=7, column=col_idx + 1).fill = LIGHT_BLUE_FILL

        # ── Key insight callout ──
        row = 9
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        insight = ws.cell(row=row, column=1)
        insight.value = (
            "КЛЮЧЕВОЙ ВЫВОД: Средняя конкуренция на тендер составляет всего "
            f"{mkt_12m.get('avg_participants', 0)} участника. "
            "Большинство крупных тендеров выигрывают компании с низким рейтингом (DDD). "
            "Для компании с рейтингом B — это окно возможностей."
        )
        insight.font = Font(name="Calibri", size=10, bold=True, color=GREEN_FONT)
        insight.fill = GREEN_FILL
        insight.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 40

        # ── Monthly trend ──
        row = 11
        ws.cell(row=row, column=1, value="Ежемесячная динамика (12 мес.)").font = SUBTITLE_FONT
        row += 1
        if not monthly.empty:
            row = self._write_df(ws, monthly, start_row=row)

        # ── Regional distribution ──
        row += 2
        ws.cell(row=row, column=1, value="Распределение по регионам").font = SUBTITLE_FONT
        row += 1
        if not regional.empty:
            row = self._write_df(ws, regional, start_row=row)

        # ── Top customers ──
        row += 2
        ws.cell(row=row, column=1, value="Топ-10 крупнейших заказчиков (12 мес.)").font = SUBTITLE_FONT
        row += 1
        if not top_cust.empty:
            row = self._write_df(ws, top_cust, start_row=row)

        # ── Rating distribution ──
        row += 2
        ws.cell(row=row, column=1, value="Распределение компаний по рейтингу").font = SUBTITLE_FONT
        row += 1
        if not rating_dist.empty:
            end = self._write_df(ws, rating_dist, start_row=row)
            # Color-code rating column
            for r in range(row + 1, end + 1):
                cell = ws.cell(row=r, column=1)
                rating = str(cell.value or "")
                color = RATING_COLORS.get(rating)
                if color:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        self._auto_fit(ws)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ═══════════════════════════════════════════════════════════
    #  SHEET 2: ТОП-15 КОМПАНИЙ
    # ═══════════════════════════════════════════════════════════

    async def _sheet_top_companies(self, wb: Workbook, uet_stir: str) -> None:
        ws = wb.create_sheet("Топ-15 компаний")
        df = await get_top_companies_overall(self.pool, limit=15)

        ws.merge_cells("A1:J1")
        ws["A1"] = "ТОП-15 СТРОИТЕЛЬНЫХ КОМПАНИЙ ПО КОЛИЧЕСТВУ ПОБЕД НА ТЕНДЕРАХ"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=NAVY)

        ws.merge_cells("A2:J2")
        ws["A2"] = "Период: последние 12 месяцев | Источник: etender.uzex.uz + reyting.mc.uz"
        ws["A2"].font = SMALL_FONT

        if df.empty:
            ws["A4"] = "Нет данных"
            return

        end_row = self._write_df(ws, df, start_row=4)

        # Highlight Tashkent companies and color-code ratings
        stir_col = None
        rating_col = None
        region_col = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == "СТИР":
                stir_col = col_idx
            elif col_name == "Рейтинг":
                rating_col = col_idx
            elif col_name == "Регион":
                region_col = col_idx

        for row_idx in range(5, end_row + 1):
            # Color ratings
            if rating_col:
                cell = ws.cell(row=row_idx, column=rating_col)
                rating = str(cell.value or "")
                color = RATING_COLORS.get(rating)
                if color:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Bold Tashkent rows
            if region_col:
                region_cell = ws.cell(row=row_idx, column=region_col)
                region_val = str(region_cell.value or "").lower()
                if "toshkent" in region_val or "тошкент" in region_val or "ташкент" in region_val:
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=c).font = Font(
                            name="Calibri", size=10, bold=True
                        )

            # Highlight UET row
            if stir_col:
                if str(ws.cell(row=row_idx, column=stir_col).value) == uet_stir:
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=c).fill = UET_ROW_FILL

        # Insight box below table
        row = end_row + 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        note = ws.cell(row=row, column=1)
        note.value = (
            "ОБРАТИТЕ ВНИМАНИЕ: Большинство лидеров по тендерам имеют низкий рейтинг (DDD/DD). "
            "Рейтинг не является барьером для участия в тендерах. "
            "UET Construction с рейтингом B может использовать это как конкурентное преимущество."
        )
        note.font = Font(name="Calibri", size=10, bold=True, color=YELLOW_FONT)
        note.fill = YELLOW_FILL
        note.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 40

        self._auto_fit(ws)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ═══════════════════════════════════════════════════════════
    #  SHEET 3: ПРОФИЛЬ UET (The Key Sheet)
    # ═══════════════════════════════════════════════════════════

    async def _sheet_uet_profile(self, wb: Workbook, stir: str) -> None:
        ws = wb.create_sheet("Профиль UET")

        uet = await get_uet_profile(self.pool, stir)
        if not uet:
            ws["A1"] = f"Компания со СТИР {stir} не найдена"
            return

        score = float(uet.get("rating_score") or 0)
        percentile = await get_uet_rating_percentile(self.pool, Decimal(str(score)))
        total_rated = percentile.get("total_rated", 1)
        rank_position = percentile.get("at_or_above", 0)
        pct = round((1 - rank_position / total_rated) * 100, 1) if total_rated else 0

        breakdown = await get_uet_rating_breakdown(self.pool, stir)
        benchmark = await get_top50_benchmark(self.pool)
        detail = await get_uet_rating_detail(self.pool, stir)
        gaps = await get_uet_competitiveness_gaps(self.pool, stir)
        mkt_12m = await get_market_summary_12m(self.pool)

        # ── Title ──
        ws.merge_cells("A1:H1")
        ws["A1"] = f"UET CONSTRUCTION — СТРАТЕГИЧЕСКИЙ ПРОФИЛЬ"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

        ws.merge_cells("A2:H2")
        ws["A2"] = f"СТИР: {stir} | Регион: {uet.get('region', 'N/A')} | Рейтинг: {uet.get('rating_letter', 'N/A')} ({score} баллов)"
        ws["A2"].font = Font(name="Calibri", size=11, color="333333")

        # ═══ SECTION: STRENGTHS (green) ═══
        row = 4
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        s = ws.cell(row=row, column=1, value="СИЛЬНЫЕ СТОРОНЫ")
        s.font = Font(name="Calibri", size=13, bold=True, color=GREEN_FONT)
        s.fill = GREEN_FILL
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = GREEN_FILL

        row += 1
        strengths = [
            (
                f"Рейтинг B — топ {round(rank_position / total_rated * 100, 1)}% из {total_rated:,} компаний",
                f"Только {rank_position} компаний в Узбекистане имеют такой же или более высокий рейтинг. "
                f"88.7% строительных компаний имеют рейтинг DDD или ниже."
            ),
            (
                f"{uet.get('employee_count') or 391} сотрудников, {uet.get('specialist_count', 14)} специалистов",
                "Крупный штат с квалифицированными кадрами. "
                "Рейтинговый балл за кадры: 12.24 из 18.40 (66.5%)."
            ),
            (
                "Сертификация ISO 9001, ISO 14001, ISO 45001",
                "Полный набор международных сертификатов. "
                "Большинство конкурентов на etender не имеют ни одного."
            ),
            (
                "Сильные финансовые показатели: 61.9% от максимума",
                "UET опережает средний показатель Топ-50 компаний (44.5%) по финансовым "
                "индикаторам на 17.4 п.п. Это объективное преимущество в тендерах."
            ),
        ]
        for title, desc in strengths:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
            c1 = ws.cell(row=row, column=1, value=title)
            c1.font = Font(name="Calibri", size=10, bold=True, color=GREEN_FONT)
            c2 = ws.cell(row=row, column=4, value=desc)
            c2.font = DATA_FONT
            c2.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 35
            row += 1

        # ═══ SECTION: GAPS (red) ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        g = ws.cell(row=row, column=1, value="КЛЮЧЕВЫЕ РАЗРЫВЫ")
        g.font = Font(name="Calibri", size=13, bold=True, color=RED_FONT)
        g.fill = RED_FILL
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = RED_FILL

        row += 1
        gap_items = [
            (
                "0 побед на etender.uzex.uz",
                "UET полностью отсутствует на крупнейшей публичной площадке госзакупок. "
                "Конкуренты, включая SHAFOAT QURILISH (тоже B-рейтинг), уже выигрывают крупные тендеры. "
                "UET невидим для государственных заказчиков, использующих etender."
            ),
            (
                "Конкурентоспособность: 6.72 из 54.10 (12.4%)",
                "Категория 'Конкурентоспособность' — это техника и оборудование. "
                "UET набирает 12.4%, тогда как Топ-50 компаний в среднем 36.9%. "
                "Это крупнейший источник потерянных рейтинговых баллов (~47 из 99 возможных)."
            ),
            (
                "Недостаток тяжёлой строительной техники",
                "0 бульдозеров, 0 катков, 0 фронтальных погрузчиков, 0 бетононасосов. "
                "Каждая единица техники — это дополнительные рейтинговые баллы."
            ),
        ]
        for title, desc in gap_items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
            c1 = ws.cell(row=row, column=1, value=title)
            c1.font = Font(name="Calibri", size=10, bold=True, color=RED_FONT)
            c2 = ws.cell(row=row, column=4, value=desc)
            c2.font = DATA_FONT
            c2.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 45
            row += 1

        # ═══ SECTION: OPPORTUNITY (yellow) ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        o = ws.cell(row=row, column=1, value="ВОЗМОЖНОСТИ")
        o.font = Font(name="Calibri", size=13, bold=True, color=YELLOW_FONT)
        o.fill = YELLOW_FILL
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = YELLOW_FILL

        mkt_vol = mkt_12m.get("total_volume", 0)
        one_pct = int(mkt_vol * 0.01) if mkt_vol else 0

        row += 1
        opp_items = [
            (
                f"1% рынка = {_fmt_uzs(one_pct)}",
                f"Объём рынка строительных тендеров за 12 месяцев: {_fmt_uzs(mkt_vol)}. "
                f"Если UET захватит хотя бы 1% — это {_fmt_uzs(one_pct)} нового дохода "
                "через публичные закупки."
            ),
            (
                "Слабая конкуренция на тендерах",
                f"Среднее число участников тендера — всего {mkt_12m.get('avg_participants', 3)}. "
                "Многие крупные тендеры в Ташкенте проходят с 2-3 участниками. "
                "Барьер входа минимальный."
            ),
            (
                "Рейтинг как дифференциатор",
                "92% компаний на etender имеют рейтинг ниже B (DDD/DD/D) или вообще без рейтинга. "
                "Рейтинг B при правильном позиционировании — аргумент для заказчиков, "
                "выбирающих надёжного подрядчика."
            ),
        ]
        for title, desc in opp_items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
            c1 = ws.cell(row=row, column=1, value=title)
            c1.font = Font(name="Calibri", size=10, bold=True, color=YELLOW_FONT)
            c2 = ws.cell(row=row, column=4, value=desc)
            c2.font = DATA_FONT
            c2.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 45
            row += 1

        # ═══ SECTION: Rating breakdown vs benchmark ═══
        row += 2
        ws.cell(row=row, column=1, value="РЕЙТИНГ: UET vs ТОП-50 КОМПАНИЙ").font = SUBTITLE_FONT
        row += 1

        if not breakdown.empty and not benchmark.empty:
            # Merge the two DataFrames
            merged = breakdown.merge(
                benchmark[["Категория", "Топ-50 ср. баллы", "Топ-50 %"]],
                on="Категория", how="left",
            )
            merged["Разница"] = merged["UET %"] - merged["Топ-50 %"]
            end_row = self._write_df(ws, merged, start_row=row)

            # Color-code the difference column
            diff_col = len(merged.columns)
            for r in range(row + 1, end_row + 1):
                cell = ws.cell(row=r, column=diff_col)
                try:
                    val = float(cell.value) if cell.value else 0
                except (ValueError, TypeError):
                    val = 0
                if val > 0:
                    cell.fill = GREEN_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_FONT)
                elif val < -5:
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True, color=RED_FONT)
            row = end_row

        # ═══ Competitiveness gaps detail ═══
        row += 2
        ws.cell(row=row, column=1, value="ПОТЕНЦИАЛ РОСТА РЕЙТИНГА: Конкурентоспособность").font = SUBTITLE_FONT
        ws.cell(row=row + 1, column=1, value="Показатели, где UET теряет больше всего баллов").font = SMALL_FONT
        row += 2

        if not gaps.empty:
            top_gaps = gaps.head(10)
            end_row = self._write_df(ws, top_gaps, start_row=row)
            # Color potential column
            pot_col = len(top_gaps.columns)
            for r in range(row + 1, end_row + 1):
                cell = ws.cell(row=r, column=pot_col)
                try:
                    val = float(cell.value) if cell.value else 0
                except (ValueError, TypeError):
                    val = 0
                if val > 1:
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True, color=RED_FONT)
            row = end_row

        self._auto_fit(ws)
        # Adjust column A to be wider for labels
        ws.column_dimensions["A"].width = 40
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ═══════════════════════════════════════════════════════════
    #  SHEET 4: КОНКУРЕНТЫ (Competitor Analysis)
    # ═══════════════════════════════════════════════════════════

    async def _sheet_competitors(
        self, wb: Workbook, uet_stir: str, compare_stirs: list[str] | None,
    ) -> None:
        ws = wb.create_sheet("Конкуренты")

        tashkent_comp = await get_tashkent_competitors(self.pool, limit=15)

        # ── Title ──
        ws.merge_cells("A1:J1")
        ws["A1"] = "КОНКУРЕНТНЫЙ АНАЛИЗ: ТАШКЕНТСКИЙ РЫНОК"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

        ws.merge_cells("A2:J2")
        ws["A2"] = "Компании, наиболее активные на строительных тендерах в Ташкенте"
        ws["A2"].font = SMALL_FONT

        # ── Tashkent competitors table ──
        row = 4
        ws.cell(row=row, column=1, value="Топ-15 конкурентов в Ташкенте (по кол-ву побед)").font = SUBTITLE_FONT
        row += 1

        if not tashkent_comp.empty:
            end_row = self._write_df(ws, tashkent_comp, start_row=row)

            # Color-code ratings
            rating_col = None
            for ci, cn in enumerate(tashkent_comp.columns, 1):
                if cn == "Рейтинг":
                    rating_col = ci
                    break
            if rating_col:
                for r in range(row + 1, end_row + 1):
                    cell = ws.cell(row=r, column=rating_col)
                    rating = str(cell.value or "")
                    color = RATING_COLORS.get(rating)
                    if color:
                        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            row = end_row + 1

        # Insight box
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ins = ws.cell(row=row, column=1)
        ins.value = (
            "ВЫВОД: Большинство активных конкурентов в Ташкенте не имеют рейтинга или имеют "
            "рейтинг DDD-CC. UET с рейтингом B (35.36) объективно превосходит по качеству "
            "практически всех активных участников тендеров в регионе."
        )
        ins.font = Font(name="Calibri", size=10, bold=True, color=GREEN_FONT)
        ins.fill = GREEN_FILL
        ins.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 40

        # ── Side-by-side comparison ──
        if compare_stirs:
            row += 2
            ws.cell(row=row, column=1, value="UET vs Выбранные конкуренты").font = SUBTITLE_FONT
            row += 1

            peer_df = await get_peer_comparison(self.pool, uet_stir, compare_stirs)
            if not peer_df.empty:
                end_row = self._write_df(ws, peer_df, start_row=row)
                # Highlight UET row
                for r in range(row + 1, end_row + 1):
                    stir_col = None
                    for ci, cn in enumerate(peer_df.columns, 1):
                        if cn == "СТИР":
                            stir_col = ci
                            break
                    if stir_col and str(ws.cell(row=r, column=stir_col).value) == uet_stir:
                        for c in range(1, len(peer_df.columns) + 1):
                            ws.cell(row=r, column=c).fill = UET_ROW_FILL
                row = end_row

            # Rating category comparison
            row += 2
            ws.cell(row=row, column=1, value="Сравнение рейтинговых категорий").font = SUBTITLE_FONT
            row += 1

            all_stirs = [uet_stir] + [s for s in compare_stirs if s != uet_stir]
            rating_comp = await get_peer_rating_comparison(self.pool, all_stirs)
            if not rating_comp.empty:
                self._write_df(ws, rating_comp.reset_index(), start_row=row)

        self._auto_fit(ws)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ═══════════════════════════════════════════════════════════
    #  SHEET 5: РЕКОМЕНДАЦИИ (Recommendations)
    # ═══════════════════════════════════════════════════════════

    async def _sheet_recommendations(self, wb: Workbook, uet_stir: str) -> None:
        ws = wb.create_sheet("Рекомендации")

        big_tenders = await get_big_tashkent_tenders(self.pool, limit=15)
        tash_customers = await get_tashkent_customers(self.pool, limit=10)
        gaps = await get_uet_competitiveness_gaps(self.pool, uet_stir)
        mkt_12m = await get_market_summary_12m(self.pool)

        # ── Title ──
        ws.merge_cells("A1:H1")
        ws["A1"] = "РЕКОМЕНДАЦИИ ДЛЯ UET CONSTRUCTION"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=NAVY)

        ws.merge_cells("A2:H2")
        ws["A2"] = "Стратегические действия на основе анализа рынка"
        ws["A2"].font = SMALL_FONT

        # ═══ RECOMMENDATION 1: Enter etender ═══
        row = 4
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        r1 = ws.cell(row=row, column=1)
        r1.value = "1. ВЫХОД НА ETENDER.UZEX.UZ — ПРИОРИТЕТ №1"
        r1.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        r1.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        row += 1
        rec1_items = [
            "Зарегистрироваться на etender.uzex.uz как поставщик строительных услуг",
            "Настроить мониторинг новых тендеров по категориям строительства в Ташкенте",
            "Начать с тендеров среднего размера (500 млн — 2 млрд UZS) для накопления опыта",
            "Использовать рейтинг B как конкурентное преимущество в тендерной документации",
            "Пиковый сезон тендеров: апрель — сентябрь. Подготовку начинать в марте.",
        ]
        for item in rec1_items:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            c = ws.cell(row=row, column=1, value=f"  {item}")
            c.font = DATA_FONT
            c.alignment = Alignment(wrap_text=True)
            row += 1

        # ═══ RECOMMENDATION 2: Target customers ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        r2 = ws.cell(row=row, column=1)
        r2.value = "2. ЦЕЛЕВЫЕ ЗАКАЗЧИКИ В ТАШКЕНТЕ"
        r2.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        r2.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1, value="Крупнейшие заказчики строительных работ в Ташкенте (за 12 мес.)").font = SECTION_FONT
        row += 1
        if not tash_customers.empty:
            end_row = self._write_df(ws, tash_customers, start_row=row)
            row = end_row + 1

        # ═══ RECOMMENDATION 3: Tenders to watch ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        r3 = ws.cell(row=row, column=1)
        r3.value = "3. КРУПНЫЕ ТЕНДЕРЫ В ТАШКЕНТЕ (последние 6 мес.) — УПУЩЕННЫЕ ВОЗМОЖНОСТИ"
        r3.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        r3.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(
            row=row, column=1,
            value="Эти тендеры прошли без участия UET. Большинство победителей имеют рейтинг ниже UET."
        ).font = SMALL_FONT
        row += 1

        if not big_tenders.empty:
            end_row = self._write_df(ws, big_tenders, start_row=row)
            # Color-code winner ratings
            rating_col = None
            for ci, cn in enumerate(big_tenders.columns, 1):
                if cn == "Рейтинг победителя":
                    rating_col = ci
                    break
            if rating_col:
                for r in range(row + 1, end_row + 1):
                    cell = ws.cell(row=r, column=rating_col)
                    rv = str(cell.value or "")
                    color = RATING_COLORS.get(rv)
                    if color:
                        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    elif not rv or rv == "None" or rv == "":
                        cell.value = "—"
            row = end_row + 1

        # ═══ RECOMMENDATION 4: Improve rating ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        r4 = ws.cell(row=row, column=1)
        r4.value = "4. ПЛАН ПОВЫШЕНИЯ РЕЙТИНГА: ОТ B К BB И ВЫШЕ"
        r4.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        r4.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        gap_to_bb = 40.0 - score if (score := float((await get_uet_profile(self.pool, uet_stir)).get("rating_score") or 0)) else 5
        ws.cell(
            row=row, column=1,
            value=f"UET: {score} баллов. До BB (40.0) не хватает {gap_to_bb:.1f} баллов. "
                  f"Основной источник роста — категория «Конкурентоспособность» (техника)."
        ).font = Font(name="Calibri", size=10, bold=True, color=YELLOW_FONT)
        ws.cell(row=row, column=1).fill = YELLOW_FILL
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = YELLOW_FILL
        ws.row_dimensions[row].height = 35
        row += 2

        ws.cell(row=row, column=1, value="Топ-10 показателей, где UET теряет баллы (потенциал роста)").font = SECTION_FONT
        row += 1
        if not gaps.empty:
            top_gaps = gaps.head(10)
            end_row = self._write_df(ws, top_gaps, start_row=row)
            pot_col = len(top_gaps.columns)
            for r in range(row + 1, end_row + 1):
                cell = ws.cell(row=r, column=pot_col)
                try:
                    val = float(cell.value) if cell.value else 0
                except (ValueError, TypeError):
                    val = 0
                if val >= 2:
                    cell.fill = RED_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True, color=RED_FONT)
                elif val >= 1:
                    cell.fill = YELLOW_FILL
            row = end_row + 1

        # ═══ RECOMMENDATION 5: Seasonal strategy ═══
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        r5 = ws.cell(row=row, column=1)
        r5.value = "5. СЕЗОННАЯ СТРАТЕГИЯ"
        r5.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
        r5.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        for c in range(2, 9):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

        row += 1
        seasonal = [
            "Март: Подготовка тендерной документации, анализ предстоящих закупок",
            "Апрель — Июнь: Пик публикации тендеров. Активное участие.",
            "Июль: Исторический максимум объёма (1.57 трлн UZS в июле 2025). Пиковая нагрузка.",
            "Август — Сентябрь: Второй пик активности. Завершение летних проектов.",
            "Октябрь — Декабрь: Объёмы снижаются. Время для развития мощностей и техники.",
            "Январь — Февраль: Планирование на следующий сезон.",
        ]
        for item in seasonal:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws.cell(row=row, column=1, value=f"  {item}").font = DATA_FONT
            row += 1

        self._auto_fit(ws)
        ws.column_dimensions["A"].width = 40
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1

    # ═══════════════════════════════════════════════════════════
    #  HELPERS
    # ═══════════════════════════════════════════════════════════

    def _write_df(
        self,
        ws: Any,
        df: pd.DataFrame,
        start_row: int = 1,
    ) -> int:
        """Write a DataFrame with styled headers. Returns last data row."""
        if df.empty:
            return start_row

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        for row_offset, (_, row_data) in enumerate(df.iterrows(), start=1):
            data_row = start_row + row_offset
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=data_row, column=col_idx)
                if pd.isna(value):
                    cell.value = ""
                elif hasattr(value, "item"):
                    cell.value = value.item()
                else:
                    cell.value = value

                cell.font = DATA_FONT
                cell.border = THIN_BORDER

                if row_offset % 2 == 0:
                    cell.fill = ALT_ROW_FILL

                if isinstance(cell.value, (int, float)):
                    if abs(cell.value) >= 1000:
                        cell.number_format = CURRENCY_FORMAT

        return start_row + len(df)

    def _auto_fit(self, ws: Any) -> None:
        """Auto-fit column widths based on content."""
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    length = len(str(cell.value))
                    max_length = max(max_length, length)
            adjusted = min(max_length + 3, 55)
            ws.column_dimensions[col_letter].width = max(adjusted, 10)
